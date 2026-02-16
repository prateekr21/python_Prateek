import openpyxl

path="C:\\Users\\hpx\\Desktop\\VIN\\testdata.xlsx"
workbook=openpyxl.load_workbook(path)

sheet=workbook.active #get active sheet

sheet.cell(1,1).value="123"
sheet.cell(1,2).value="smit"
sheet.cell(1,3).value="QA"

sheet.cell(2,1).value="124"
sheet.cell(2,2).value="spiderman"
sheet.cell(2,3).value="Dev"

sheet.cell(3,1).value="125"
sheet.cell(3,2).value="ironman"
sheet.cell(3,3).value="Admin"

sheet.cell(4,1).value="126"
sheet.cell(4,2).value="hulk"
sheet.cell(4,3).value="Tester"


workbook.save(path)