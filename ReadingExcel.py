import openpyxl

file="C:\\Users\\hpx\\Desktop\\VIN\\data.xlsx"
# File—-----> Workbook—---->sheets—->Rows—--->cells
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"] 
row=sheet.max_row #number of rows
col=sheet.max_column #number of columns

#read data from cell
for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value,end="     ")

    print()
    