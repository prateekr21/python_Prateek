import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row

def getColumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column

def readData(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rownum,colnum).value

def writeData(file,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rownum,colnum).value=data
    workbook.save(file)

def fillGreenColor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenFill=PatternFill(start_color="00FF00",end_color="00FF00",fill_type="solid")
    sheet.cell(rownum,colnum).fill=greenFill
    workbook.save(file)

def fillRedColor(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redFill=PatternFill(start_color="FF0000",end_color="FF0000",fill_type="solid")
    sheet.cell(rownum,colnum).fill=redFill
    
    workbook.save(file)